import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

base = r'G:\공유 드라이브\02. 소상공인 법률자문 및 채무조정 용역\2025_소상공인 법률지원 용역\2. 채무조정'

targets = [
    ('2-3. 개인회생', '채무조정_16928_이효은[완]'),
    ('2-3. 개인회생', '채무조정_17007_김진우[완]'),
    ('2-3. 개인회생', '채무조정_17022_최은서[완]'),
    ('2-2. 개인파산', '채무조정_16932_박수영[완]'),
    ('2-2. 개인파산', '채무조정_16960_김선희[완]'),
    ('2-5. 새출발기금', '채무조정_16957_김명옥[완]'),
    ('2-5. 새출발기금', '채무조정_16946_김태현[완]'),
    ('2-5. 새출발기금', '채무조정_17154_지심결[완]'),
    ('2-4. 신용회복위', '채무조정_16935_김향경[완]'),
    ('2-4. 신용회복위', '채무조정_17101_김윤환[완]'),
]

TEMPLATE_VALS = {'소득의 종류', '금액(만원)', '직업 및 소득원', '혼인관계', '`', '비고'}

def clean(val):
    if val is None:
        return ''
    s = str(val).strip()
    if s in TEMPLATE_VALS:
        return ''
    return s

def to_num(val):
    if val is None:
        return 0
    if str(val).strip() in ('', '`', '금액(만원)'):
        return 0
    try:
        return int(float(val))
    except:
        return 0

for sub, folder in targets:
    fpath = os.path.join(base, sub, folder)
    if not os.path.exists(fpath):
        print(f'--- {folder}: NOT FOUND ---')
        continue
    xlsx_files = [f for f in os.listdir(fpath) if f.endswith('.xlsx')]
    if not xlsx_files:
        print(f'--- {folder}: NO XLSX ---')
        continue
    xlsx = os.path.join(fpath, xlsx_files[0])
    try:
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        if '상담일지' not in wb.sheetnames:
            print(f'--- {folder}: NO 상담일지 sheet ---')
            continue
        ws = wb['상담일지']

        # dump all non-empty cells for first file to understand structure
        name = clean(ws['D5'].value)
        phone = clean(ws['H7'].value)
        birth_raw = clean(ws['I5'].value)
        age = to_num(ws['H5'].value)
        addr = clean(ws['D9'].value)
        biz_status = clean(ws['D10'].value)

        # debt
        credit = to_num(ws['D22'].value)
        secured = to_num(ws['D23'].value)
        tax = to_num(ws['D24'].value)
        private = to_num(ws['D25'].value)
        goods = to_num(ws['D26'].value)
        etc_debt = to_num(ws['D27'].value)
        total_debt = to_num(ws['D28'].value)

        # assets
        deposit = to_num(ws['H22'].value)
        total_assets = to_num(ws['H28'].value)

        # income - try multiple rows
        income1 = to_num(ws['C37'].value)
        income2 = to_num(ws['C38'].value)
        job1 = clean(ws['E36'].value)
        job2 = clean(ws['E37'].value)
        income_type = clean(ws['A37'].value)

        # family
        marriage = clean(ws['B40'].value)
        dependents = to_num(ws['E40'].value)
        minors = to_num(ws['G40'].value)
        elderly = to_num(ws['I40'].value)

        # special
        gambling = clean(ws['B43'].value)
        prev_case = clean(ws['B44'].value)
        health = clean(ws['B45'].value)
        etc_note = clean(ws['B46'].value)

        print(f'=== {folder} ({sub}) ===')
        print(f'  이름: {name}')
        print(f'  전화: {phone}')
        print(f'  생년: {birth_raw} (나이: {age})')
        print(f'  주소: {addr}')
        print(f'  사업: {biz_status}')
        print(f'  채무: 신용={credit} 담보={secured} 세금={tax} 사채={private} 물품={goods} 기타={etc_debt}')
        print(f'  채무총액: {total_debt}만원 | 재산: {total_assets}만원 (보증금:{deposit})')
        print(f'  소득: {income1}만원 | 종류: {income_type} | 직업: {job1} {job2}')
        print(f'  가족: {marriage} 부양{dependents} 미성년{minors} 65세+{elderly}')
        print(f'  특기: 도박={gambling} 전력={prev_case} 치료={health} 기타={etc_note}')
        print()
    except Exception as e:
        print(f'--- {folder}: ERROR {e} ---')
        import traceback
        traceback.print_exc()
        print()
